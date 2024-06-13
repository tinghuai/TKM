# coding=UTF-8



import csv
import json
from openpyxl import load_workbook
import os
import re
import emoji
import time
import os
import shutil

def setDir(filepath):
    '''
    如果文件夹不存在就创建，如果文件存在就清空！
    :param filepath:需要创建的文件夹路径
    :return:
    '''
    if not os.path.exists(filepath):
        os.mkdir(filepath)
    else:
        shutil.rmtree(filepath)
        os.mkdir(filepath)

# 如果要多线程,外置的flag要写多个,可以ins一个,tw一个
# csv存储会有各种问题，直接改为txt按行存储

def read_excel(excel_name, sheet_name, column, row_s, row_f):
    """
    读取excel文件
    :param excel_name:文件名,可以是路径
    :param sheet_name:dheet名称
    :param column:读取的列,用"A" "B"...表示
    :param row_s:开始行,从1开始,读取到excel实际对应的row行
    :param row_f:结束行,读取到excel实际对应的row行
    :return:返回那一列到row行的一个列表
    """
    # 1.打开 Excel 表格并获取表格名称
    workbook = load_workbook(filename=excel_name)
    # print(workbook.sheetnames)
    # 2.通过 sheet 名称获取表格
    sheet = workbook[sheet_name]
    # print(sheet)
    # 3.获取表格的尺寸大小(几行几列数据) 这里所说的尺寸大小，指的是 excel 表格中的数据有几行几列，针对的是不同的 sheet 而言。
    # print(sheet.dimensions)
    # 4.获取表格内某个格子的数据
    # 1 sheet["A1"]方式
    # cell = sheet["A1"]
    data = []
    for i in range(row_s, row_f+1):
        cell = sheet[column + str(i)]
        data.append(cell.value)

    # print(data)
    return data

def add_to_match_list(name):
    # 判断csv是否为空
    # 如果是空的，要创建head
    #如果不是空的，

    # 读取name列中是否已经有该用户
    # 如果有，则获取name列中的索引，并且把那个值置为none（相当于废弃那一行）
    # 如果没有，则直接在结尾append，并判断结尾的所有是否是text的下一个，并且把append的索引行号返回（返回后要在txt中更新索引号）
    return

def include_other_language(content):
    '''
    识别是否包含中文（简繁） 日文 韩文 俄文，用python re包的正则表达式实现
    :param content:输入的句子，是str
    :return:返回match到的次数
    '''
    Pattern = re.compile(
        u'[\u4e00-\u9fa5\u3040-\u309F\u30A0-\u30FF\uAC00-\uD7A3\u0400-\u052f]+')  # 中文（简繁） 日文 韩文 俄文
    match = Pattern.search(content)
    return match

def use_re_get_text_and_time(s):
    '''
    使用正则表达式取出一段字符串中的文本信息和时间，并将时间转化为时间戳
    去除url
    去除单引号和双引号
    因为csv是以逗号为分隔的，所以将text中的逗号进行替换
    没有处理emoji表情，需要后续手动处理
    :param s: 需要提取的句子
    :return: 返回text和time
    '''

    # 提取text的规则
    p_draw_text = re.compile(", 'text': (.*?), 'truncated': ", re.S)  #非贪婪匹配,如果没有”？“则是贪婪匹配
    text_ori = re.findall(p_draw_text, s)
    if text_ori == []: #因为会有空行，空行识别出来是一个空列表
        return '', ''
    # 去除url链接
    text = re.sub(r'(https|http)?:\/\/(\w|\.|\/|\?|\=|\&|\%)*\b', '', text_ori[0])  # 第二位是用来替换的
    #去除前后的引号
    text = text[1:-1]
    text = text.replace('\\n', '')  # 去除句子中的“\n”,因为句子中不是换行，只是单纯的字符
    # 去除字符串中引号，不然csv读取会出问题
    text = text.replace('"', '')
    text = text.replace("'", '')
    text = text.replace('\n', '').replace('\t', '')
    text = text.replace(',', ' ')  # 防止逗号前后的两个单词连成1个，用空格替代
    text = text.replace('::', ' ')  # csv碰到这个会自动换行
    text = text.replace(':', ' ')

    p_draw_time = re.compile("_json={'created_at': '(.*?)', 'id':", re.S)
    time_ori = re.findall(p_draw_time, s)
    if time_ori == []:
        return '', ''
    time_stamp = int(time.mktime(time.strptime(time_ori[0], "%a %b %d %H:%M:%S +0000 %Y")))  # 转化格式，并直接输出时间戳
    return text, time_stamp

def check_column(filename,column_num,delimiter="\t"):
    '''
    检查输出的csv的列数是否符合要求，防止分隔符出现问题
    造这个函数的原因：excel和pycharm中自带的csv查看器都不能正常处理分隔符
    :param filename: 文件路径
    :param column: 应该有的列数
    :param delimiter: 分隔符
    :return: 如果列数不等于应该有的列数，则返回错误信息，是字符串；如果列数等于应该有的列数，则返回-1
    '''
    with open(filename, 'r', encoding='utf-8') as f:
        reader = csv.reader(f, delimiter=delimiter)
        it = -1
        for row in reader:
            it = it+1
            if len(row) != column_num:
                s ="in file" + filename + "in row " + it + "有" + len(row) + "列"
                print(s)
                return s
                break
    return 1



if __name__ == "__main__":
    # ins和tw共有的变量
    ins_tw_start = 1
    ins_tw_finish = 3000
    ins_twi_match_row = []
    ins_twi_match_rows = []

    # 创建一个ins_save_user_number的列表，并且append当前处理的，此处直接用tw_save_user_name来getindex来代替
    # 创建一个ins_save_user_name的列表，并且append当前处理的
    ins_save_user_name = []
    tw_save_user_name = []
    # 数据集所在位置
    original_data_path = "datasets/original/"

    # 先清空文件夹
    setDir("datasets/original_draw/ins")
    setDir("datasets/original_draw/tw")
    if os.path.exists("datasets/original_draw/ins_tw_match.csv") == True:
        os.remove("datasets/original_draw/ins_tw_match.csv")
    # exit(0)  # 停止运行

    # 以下都是处理ins自己的变量
    ins_loginfo = open('loginfo/draw_test_and_time_insloginfo.txt', mode='w+')  # w+覆盖读写
    # 读取ins_tw.excl文件的起始行和终止行


    #当前处理的社交网络是sn（socil network）
    sn = "ins"
    #todo 读取flag_1_ins.txt中的序号,是对应读ins_tw哪一行的序号(循环从这个序号下一个开始读)
    #todo 读取flag_2_ins.txt中的序号，为flag_2,是转化后的文件的序号(循环从这个序号下一个开始写)
    #按行读取black_list,并且保存到一个list中
    blacklist_file = original_data_path+'blacklist.txt'
    blacklist = []
    f = open(blacklist_file)
    line = f.readline()
    while line:
        if(line.strip("\n") not in blacklist):
            blacklist.append(line.strip("\n"))  # 去除每行结尾的换行符
            # print(line, end='')
        line = f.readline()
    f.close()
    # print(blacklist)

    #读取ins_tw.csv,获取ins列，并且返回一个列表
    #todo 读取的start是flag_1的值+1

    ins_list = read_excel('datasets/original/ins_twi.xlsx', "ins_twi", "A", ins_tw_start, ins_tw_finish)  # ins是序号
    #读取ins_tw.csv,获取tw列，并且返回一个列表
    tw_list = read_excel('datasets/original/ins_twi.xlsx', "ins_twi", "B", ins_tw_start, ins_tw_finish)
    #遍历这个列表（同时需要限制提取文本的用户数量）（直接写在函数里面了）
    ins_and_tw_list_point = -1 # 用来将当前insuser指向的用户对上tw的列表
    for insuser in ins_list:
        ins_filename = original_data_path + "ins/" + str(insuser)+'.csv'
        ins_and_tw_list_point=ins_and_tw_list_point+1
        print("process"+str(ins_and_tw_list_point))
        #判断ins的序号对应的csv文件是否存在
        if os.path.exists(ins_filename) == False:
            # print(ins_filename)
            #如果不存在，直接进入下一个循环
            ins_loginfo.write('ins_tw表格中行'+ str(ins_and_tw_list_point+ins_tw_start) + ': ' + ins_filename+'在ins中不存在\n')
            continue
        #如果存在
        else:
            #看名字在不在blacklist中
            tw_name =str(tw_list[ins_and_tw_list_point])
            if tw_name in blacklist:
                ins_loginfo.write('ins_tw表格中行'+ str(ins_and_tw_list_point+ins_tw_start) + ': ' + str(tw_list[ins_and_tw_list_point]) + '在blacklist中\n')
                continue
            #查看一共几行,如果只有一行,直接下一个循环
            total_lines = sum(1 for line in open(ins_filename, encoding='utf-8'))
            # print(total_lines)
            if total_lines < 200:
                ins_loginfo.write('ins_tw表格中行' + str(ins_and_tw_list_point+ins_tw_start) + ': ' + ins_filename + '小于150行\n')
                continue
            # 看ins_tw对应的文件是否存在
            tw_file_name = original_data_path + "/tw/" + tw_name + '.txt'
            if os.path.exists(tw_file_name) == False:
                # 如果不存在,直接下一个循环
                ins_loginfo.write('ins_tw表格中行' + str(ins_and_tw_list_point+ins_tw_start) + ': ' + str(insuser) + '对应的tw文件'+ tw_name + '不存在\n')
                continue
            # 按行读取csv
            delete_lines = 0
            with open(ins_filename, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                # 先读取所有行，计数出现了多少行其他文字，然后算保留了多少，如果保留少于36行，则直接不看这个用户

                for row in reader:
                    text_content = row[2:-3]
                    total_temp = 0
                    for i in text_content:
                        if i !='':
                            total_temp = total_temp + len(i)
                    for cell in text_content:
                        if include_other_language(str(cell)) or total_temp>300:  # 太长的也不要
                            delete_lines = delete_lines+1
                            continue
                # print(delete_lines, end=" ")

            # 如果删除要删除的行之后，剩余的行小于50行，则这个也是一个废文件
            if (total_lines - delete_lines) < 50:
                # print(total_lines,delete_lines)
                continue

            with open(ins_filename, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                # 记录当前读到第几行
                row_index = -1

                content_rows = []  # 所有的行数，是content_row的总和
                for row in reader:
                    if len(row)<6:
                        continue
                    content_row = []  # 单行的内容
                    row_index = row_index + 1
                    #跳过是表头的第一行
                    if row_index == 0:
                        continue
                    #第一行读取用户名，并保存到临时变量中
                    if row_index == 1:
                        ins_user_name = row[1]
                        # ins_user_name = row[2]
                        # print(ins_user_name)
                    # 第二行开始
                    else:

                        #todo 写csv的行号+1
                        #读取索引的倒数第二位，保存至临时变量中，应该是str（因为要判断里面有没有空格）
                        content_time = row[-3]

                        content_time.replace(" ", "") #  去除空格
                            #判断长度和是否为纯数字，长度应为10（因为爬取时间普遍在2018年）
                                #如果不是，则直接下一行
                        # print(content_time)
                        if not content_time.isdigit():
                            continue
                        # else
                            # 如果是，
                            # todo flag_2 ++
                            ## 直接在处理文本那里统一把time append进content_row
                            ## 写入csv的当前行的时间列
                        #读取csv的第三位至倒数第三位(切片)、
                        context_test_original = row[2:-3] # 这里是一个数组的形式还需要下面拼接的操作
                        # print(context_test_original)
                        # 拼接,注意中间要加空格
                        context_text = ""
                        for cell in context_test_original:
                            context_text = context_text+" "+cell
                        # print("len :"+str(len(context_text)))
                        # print(context_text)
                        # 如果包含其他的语言，则直接continue
                        if include_other_language(context_text) or len(context_text) > 300:
                            # print("no!")
                            continue
                        # 如果语句是空的，或者只有一个空格，也直接continue
                        if context_text == "" or context_text== " ":
                            continue
                        #emoji转成文字
                        else:
                            context_text = emoji.demojize(context_text)
                        # print(context_text)
                        # 写入当前列的txt列
                        #去除url链接
                        context_text = re.sub(r'(https|http)?:\/\/(\w|\.|\/|\?|\=|\&|\%)*\b', '', context_text)
                        #去除字符串中引号，不然csv读取会出问题
                        context_text = context_text.replace('"', '')
                        context_text = context_text.replace("'", '')
                        context_text = context_text.replace('\n ', '').replace('\t', '')
                        context_text = context_text.replace(',', ' ')  # 防止逗号前后的两个单词连成1个，用空格替代
                        context_text = context_text.replace('::', ' ')  # csv碰到这个会自动换行
                        context_text = context_text.replace(':', ' ')
                        context_text = context_text.replace('”', ' ')
                        context_text = context_text.replace('‘', ' ')
                        if context_text[0] == " ":
                            context_text = context_text[1:]
                        # print(context_text)
                        content_row.append(context_text)
                        content_row.append(content_time)
                        content_rows.append(content_row)
            length_temp = len(content_rows)

            # print(length_temp)
            if len(content_rows) < 151:  # 筛选出150个帖子以上的用户
                continue
            ## 创建一个csv文件，命名为用户名的序号.csv,初始化表头
            headers = ['text', 'time']
            ins_twi_match_row = []
            with open('datasets/original_draw/ins/'+ins_user_name+'.csv', 'w+', encoding='utf8', newline='') as f3:
                writer = csv.writer(f3,delimiter="\t")
                writer.writerow(headers)
                writer.writerows(content_rows)
                # add_to_match_list(name,flag_2,username)
                ins_twi_match_row.append(ins_user_name)
                # todo 这里先appendtw的名字，如果tw哪里的文件有问题，直接在tw_save_user_name中getindex后把对应的tw名字置为none
                ins_twi_match_row.append(tw_name)
                # print(ins_twi_match_row)

                if(ins_user_name.find(',')==-1 and tw_name.find(',')==-1):
                    # 如果用户名中有逗号，则直接不要
                    ins_twi_match_rows.append(ins_twi_match_row)

                    # 创建一个ins_save_user_number的列表，并且append当前处理的，此处直接用tw_save_user_name来getindex来代替
                    # 创建一个ins_save_user_name的列表，并且append当前处理的
                    ins_save_user_name.append(ins_user_name)
                    # 创建一个tw_save_user_name,此处用来tw处理时getindex
                    tw_save_user_name.append(tw_name)
                    # 上面两个的对应关系为了用于处理twitter的数据集的时候可以对应
                    # todo flag_2更新文件内行号
                print(check_column('datasets/original_draw/ins/'+ins_user_name+'.csv',2,'\t'))
        # print(ins_save_user_name)
        # print(tw_save_user_name)
        #flag_1++
        #更新flag_1中的行号


    # tw——————————————————————————————————————————————————————————————————————————————————————————————————————————————————————————————————————
    # tw——————————————————————————————————————————————————————————————————————————————————————————————————————————————————————————————————————

    # tw要额外处理的
    # 如果出现其他语言，要排除的行删除了少于36行，则只能把那个行号的位置置为none

    # 处理ins自己的变量
    tw_loginfo = open('loginfo/draw_test_and_time_twloginfo.txt', mode='w+')  # w+覆盖读写
    # 读取ins_tw.excl文件的起始行和终止行

    # 当前处理的社交网络是sn（socil network）
    sn = "tw"
    # todo 读取flag_1_ins.txt中的序号,是对应读ins_tw哪一行的序号(循环从这个序号下一个开始读)
    # todo 读取flag_2_ins.txt中的序号，为flag_2,是转化后的文件的序号(循环从这个序号下一个开始写)

    # 读取ins_tw.csv,获取ins列，并且返回一个列表
    # todo 读取的start是flag_1的值+1
    # ins_list = read_excel('datasets/original/ins_twi.xlsx', "ins_twi", "A", ins_tw_start, ins_tw_finish)  # ins是序号
    # 读取ins_tw.csv,获取tw列，并且返回一个列表
    # tw_list = read_excel('datasets/original/ins_twi.xlsx', "ins_twi", "B", ins_tw_start, ins_tw_finish)
    # 遍历这个列表（同时需要限制提取文本的用户数量）（直接写在函数里面了）

    ins_and_tw_list_point = -1  # 用来将当前insuser指向的用户对上tw的列表
    for twuser in tw_save_user_name:
        tw_filename = original_data_path + "tw/" + str(twuser) + '.txt'
        ins_and_tw_list_point = ins_and_tw_list_point + 1
        print("process"+str(ins_and_tw_list_point))
        # 判断ins的序号对应的csv文件是否存在
        if os.path.exists(tw_filename) == False:
            # print(ins_filename)
            # 如果不存在，直接进入下一个循环
            tw_loginfo.write(
                '用户' + ': ' + tw_filename + '在文件夹中不存在\n')
            #将tw和ins的用户名匹配列表中置为NoneNoneNone
            temp_index = tw_save_user_name.index(twuser)
            ins_twi_match_rows[temp_index][1] = "NoneNoneNone"
            continue
        # 如果存在
        else:
            # 查看一共几行,如果只有一行,直接下一个循环
            total_lines = sum(1 for line in open(tw_filename, encoding='utf-8'))
            # print(total_lines)
            if total_lines < 200:
                tw_loginfo.write(
                    '用户' + ': ' + tw_filename + '小于150行\n')
                # 将tw和ins的用户名匹配列表中置为NoneNoneNone
                temp_index = tw_save_user_name.index(twuser)
                ins_twi_match_rows[temp_index][1] = "NoneNoneNone"
                continue

            # 按行读取csv
            delete_lines = 0
            with open(tw_filename, 'r', encoding='utf-8') as f4:
                # 先读取所有行，计数出现了多少行其他文字或者是空行，然后算保留了多少，如果保留少于36行，则直接不看这个用户
                for row in f4:
                    text, time_stamp = use_re_get_text_and_time(row)
                    if include_other_language(text) or text==''or text[:2] == "RT" or len(text)>300: #  包括其他语言或是空行或是回复帖子
                        delete_lines = delete_lines + 1
                # print(delete_lines, end=" ")

            # 如果删除要删除的行之后，剩余的行小于50行，则这个也是一个废文件
            if (total_lines - delete_lines) < 50:
                # print(total_lines, delete_lines)
                temp_index = tw_save_user_name.index(twuser)
                ins_twi_match_rows[temp_index][1] = "NoneNoneNone"
                continue

            with open(tw_filename, 'r', encoding='utf-8') as f5:
                # 记录当前读到第几行
                row_index = -1
                content_rows = []  # 清洗后所有的行数，是content_row的总和
                for row in f5:
                    content_row = []  # 清洗后单行的内容
                    row_index = row_index + 1
                    # todo tw中还要处理时间和文本的对应关系，看行号一不一样，和取出的东西是不是空的 ，
                    # todo 去除包含RT的
                    # todo 写csv的行号+1
                    text, time_stamp = use_re_get_text_and_time(row)
                    if include_other_language(text) or text==''or text[:2] == "RT" or text==' ' or len(text)>300: #  包括其他语言或是空行或是回复帖子、
                        continue

                    # else
                    # 如果是，
                    # todo flag_2 ++
                    ## 直接在处理文本那里统一把time append进content_row
                    ##写入csv的当前行的时间列

                    # emoji转成文字
                    else:
                        text = emoji.demojize(text)
                    # 写入当前列的txt列
                    text = text.replace('::', ' ')  # csv碰到这个会自动换行
                    text = text.replace(':', ' ')

                    content_row.append(text)
                    content_row.append(time_stamp)
                    content_rows.append(content_row)
            length_temp = len(content_rows)
            # print(length_temp)
            if len(content_rows) < 151:
                temp_index = tw_save_user_name.index(twuser)
                ins_twi_match_rows[temp_index][1] = "NoneNoneNone"
                continue
            ## 创建一个csv文件，命名为用户名的序号.csv,初始化表头
            headers = ['text', 'time']
            with open('datasets/original_draw/tw/' + twuser + '.csv', 'w+', encoding='utf8',
                      newline='') as f6:
                writer = csv.writer(f6, delimiter="\t")
                writer.writerow(headers)
                writer.writerows(content_rows)
            print(check_column('datasets/original_draw/tw/' + twuser + '.csv', 2, '\t'))


    ## todo ins_tw_match.csv 这个有问题！！！！！！！！！！！！！！！！！！！！！！没办法正常分列，其他文件有的也存在这个问题
    #创建用户名对应文件ins和tw
    headers = ['ins', 'tw']
    with open('datasets/original_draw/ins_tw_match.csv', 'w+', encoding='utf8',
              newline='') as f7:
        writer = csv.writer(f7)
        writer.writerow(headers)
        for i in ins_twi_match_rows:
            # print(i)
            writer.writerow(i)
                # todo flag_2更新文件内行号
    print(check_column('datasets/original_draw/ins_tw_match.csv', 2, ','))


    with open('datasets/original_draw/ins_tw_match.csv', 'r', encoding='utf8') as f8:
        reader = csv.reader(f8)
        total_lines = 0
        value_user_peers = 0
        for row in reader:
            total_lines = total_lines +1
            if row[0]!="NoneNoneNone" and row[1]!="NoneNoneNone":
                value_user_peers = value_user_peers+1
        print("total_lines:"+str(total_lines),"value_user_peers:"+str(value_user_peers))

        # flag_1++
        # 更新flag_1中的行号

    #当前处理的社交网络是sn（socil network）
    sn = "tw"
    #读取flag_1_ins.txt中的序号,是对应读ins_tw哪一行的序号(循环从这个序号下一个开始读)
    #读取flag_2_ins.txt中的序号，为flag_2,是转化后的文件的序号(循环从这个序号下一个开始写)
    #按行读取black_list,并且保存到一个list中
    #读取ins_tw.csv,获取ins列，并且返回一个列表
    #读取ins_tw.csv,获取tw列，并且返回一个列表
    #遍历这个列表（同时需要限制提取文本的用户数量）（如果要设置断点，可以用切片操作实现）
        #判断ins的序号对应的csv文件是否存在
            #如果不存在，直接进入下一个循环
            #如果存在
                #看名字在不在blacklist中
                #查看一共几行,如果只有一行,直接下一个循环
                # 看ins_tw对应的文件是否存在
                    # 如果不存在,直接下一个循环
                #按行读取csv
                    #跳过是表头的第一行
                    #第一行读取用户名，并保存到临时变量中
                    #第二行开始
                        #写csv的行号+1
                        #读取索引的倒数第二位，保存至临时变量中，应该是str（因为要判断里面有没有空格）
                            #判断长度和是否为纯数字，长度应为10（因为爬取时间普遍在2018年）
                                #如果不是，则直接下一行
                                #如果是，
                                    # flag_2++
                                    # 创建一个csv文件，命名为flag_2的序号.csv,初始化表头
                                    # 写入csv的当前行的时间列
                        #读取csv的第三位至倒数第三位(切片)
                            #拼接,注意中间要加空格
                            #emoji转成文字
                            # 写入当前列的txt列
                        #add_to_match_list(name,flag_2,username)
                        #flag_2更新文件内行号

        #flag_1++
        #更新flag_1中的行号

